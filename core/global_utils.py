from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import os

QR_CELULAS_POR_TIPO = {
    "FO_Base": "V2",
    "FF_Base": "O1",
    "FA_Base": "K2"
}

PRINT_AREAS = {
    "FO_Base": "A1:Y63",
    "FF_Base": "A1:Q41",
    "FA_Base": "A1:N48"
}

def mes_selecionado(meses_var, mes_atual, var_atual):
    if var_atual.get():
        for mes, var in meses_var:
            if mes != mes_atual:
                var.set(False)


def inserir_qr_no_excel(template_path, destino_path, imagem_qr_path, dados={}):
    try:
        wb = load_workbook(template_path)
        ws = wb.active

        nome_template = os.path.splitext(os.path.basename(template_path))[0]

        # Inserir imagem
        celula_destino = QR_CELULAS_POR_TIPO.get(nome_template)
        if not celula_destino:
            print(f"[ERRO] Template '{nome_template}' sem célula QR definida.")
            return

        img = ExcelImage(imagem_qr_path)
        img.width = 100
        img.height = 100
        ws.add_image(img, celula_destino)

        # Definir manualmente a área de impressão, se necessário
        area = PRINT_AREAS.get(nome_template)
        if area:
            ws.print_area = area
            print(f"[INFO] Área de impressão definida: {area}")

        # Preencher células com dados (conforme o tipo)
        if nome_template == "FO_Base":
            ws["J4"] = dados.get("ano", "")
            ws["D2"] = dados.get("equipa", "")
        elif nome_template == "FF_Base":
            ws["D3"] = dados.get("ano", "")
            ws["D7"] = dados.get("equipa", "")
            ws["D5"] = dados.get("mes", "")
        elif nome_template == "FA_Base":
            ws["E6"] = dados.get("ano", "")
            ws["F4"] = dados.get("equipa", "")
            ws["H6"] = dados.get("mes", "")

        wb.save(destino_path)
        print(f"[QR INSERIDO e DADOS PREENCHIDOS] {destino_path}")

    except Exception as e:
        print(f"[ERRO] Falha ao inserir QR e preencher dados: {e}")

    except Exception as e:
        print(f"[ERRO] Falha ao inserir QR: {e}")